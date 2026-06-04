from __future__ import annotations

import base64
import difflib
import io
import json
import re
from typing import Any

from openai import OpenAI
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

from app.config import OPENAI_API_KEY, OPENAI_MODEL
from app.database import GROUP_ORDER


ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/heic", "image/heif"}
ALLOWED_TEXT_TYPES = {"text/plain", "text/csv", "application/json"}
PDF_TYPE = "application/pdf"
EXCEL_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


class RoboertaAIError(RuntimeError):
    pass


def _get_client() -> OpenAI:
    if not OPENAI_API_KEY:
        raise RoboertaAIError("OPENAI_API_KEY is not configured.")
    return OpenAI(api_key=OPENAI_API_KEY)


def _parse_json_output(output_text: str) -> dict[str, Any]:
    candidate = output_text.strip()
    if candidate.startswith("```"):
        parts = candidate.split("```")
        candidate = next((part for part in parts if "{" in part and "}" in part), candidate)
        candidate = candidate.replace("json", "", 1).strip()
    try:
        return json.loads(candidate)
    except json.JSONDecodeError as exc:
        raise RoboertaAIError(f"OpenAI returned invalid JSON: {candidate[:400]}") from exc


def _extract_excel_text(content: bytes) -> str:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                parts.append(" | ".join(values))
    return "\n".join(parts)


def _extract_pdf_text(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    pages: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        page_text = (page.extract_text() or "").strip()
        if page_text:
            pages.append(f"# Page {index}\n{page_text}")
    return "\n\n".join(pages)


def _split_pdf_into_page_files(filename: str, content: bytes) -> tuple[list[tuple[str, bytes]], int]:
    reader = PdfReader(io.BytesIO(content))
    page_files: list[tuple[str, bytes]] = []
    for index, page in enumerate(reader.pages, start=1):
        writer = PdfWriter()
        writer.add_page(page)
        buffer = io.BytesIO()
        writer.write(buffer)
        page_files.append((f"{filename}.page-{index}.pdf", buffer.getvalue()))
    return page_files, len(reader.pages)


def _build_content_parts(client: OpenAI, uploads: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    parts: list[dict[str, Any]] = []
    temporary_file_ids: list[str] = []
    for upload in uploads:
        content = upload["content"]
        content_type = upload["content_type"] or "application/octet-stream"
        filename = upload["filename"]
        if content_type in ALLOWED_IMAGE_TYPES:
            encoded = base64.b64encode(content).decode("utf-8")
            parts.append(
                {
                    "type": "input_image",
                    "image_url": f"data:{content_type};base64,{encoded}",
                }
            )
            parts.append({"type": "input_text", "text": f"Filename: {filename}"})
            continue
        if content_type == PDF_TYPE:
            page_files, page_count = _split_pdf_into_page_files(filename, content)
            parts.append(
                {
                    "type": "input_text",
                    "text": f"Filename: {filename}. Total pages: {page_count}. Parse all pages, not only the first page.",
                }
            )
            for page_filename, page_bytes in page_files:
                uploaded = client.files.create(file=(page_filename, page_bytes, content_type), purpose="user_data")
                temporary_file_ids.append(uploaded.id)
                parts.append({"type": "input_file", "file_id": uploaded.id})
                parts.append({"type": "input_text", "text": f"Attached page file: {page_filename}"})
            pdf_text = _extract_pdf_text(content)
            if pdf_text:
                parts.append({"type": "input_text", "text": f"Extracted PDF text (all pages):\n{pdf_text[:180000]}"})
            continue
        if content_type in EXCEL_TYPES:
            excel_text = _extract_excel_text(content)
            parts.append({"type": "input_text", "text": f"Filename: {filename}\n{excel_text[:50000]}"})
            continue
        if content_type in ALLOWED_TEXT_TYPES:
            text = content.decode("utf-8", errors="ignore")
            parts.append({"type": "input_text", "text": f"Filename: {filename}\n{text[:50000]}"})
            continue
        raise RoboertaAIError(f"Unsupported file type: {filename} ({content_type})")
    return parts, temporary_file_ids


def _run_json_request(system_prompt: str, instruction: str, uploads: list[dict[str, Any]]) -> dict[str, Any]:
    client = _get_client()
    parts, file_ids = _build_content_parts(client, uploads)
    try:
        response = client.responses.create(
            model=OPENAI_MODEL,
            temperature=0,
            max_output_tokens=12000,
            input=[
                {"role": "system", "content": [{"type": "input_text", "text": system_prompt}]},
                {"role": "user", "content": parts + [{"type": "input_text", "text": instruction}]},
            ],
        )
        return _parse_json_output(response.output_text)
    finally:
        for file_id in file_ids:
            try:
                client.files.delete(file_id)
            except Exception:
                continue


def _chunk_weekly_ad_uploads(uploads: list[dict[str, Any]], pages_per_chunk: int = 2) -> list[list[dict[str, Any]]]:
    chunks: list[list[dict[str, Any]]] = []
    for upload in uploads:
        content_type = upload.get("content_type") or "application/octet-stream"
        if content_type != PDF_TYPE:
            chunks.append([upload])
            continue

        page_files, _ = _split_pdf_into_page_files(str(upload.get("filename") or "weekly-ad.pdf"), upload["content"])
        page_uploads = [
            {
                "filename": page_filename,
                "content_type": PDF_TYPE,
                "content": page_bytes,
            }
            for page_filename, page_bytes in page_files
        ]
        for index in range(0, len(page_uploads), pages_per_chunk):
            chunks.append(page_uploads[index : index + pages_per_chunk])
    return chunks


def _extract_page_number_from_filename(filename: str) -> int | None:
    match = re.search(r"\.page-(\d+)\.pdf$", filename)
    if not match:
        return None
    return int(match.group(1))


def _dedupe_ad_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: dict[tuple[str, str, str, int | None], dict[str, Any]] = {}
    for item in items:
        name_key = _normalize_text(item.get("name", ""))
        size_key = _normalize_text(item.get("size_text", ""))
        price_key = _normalize_text(item.get("price_text", ""))
        page_key = item.get("page_number")
        key = (name_key, size_key, price_key, page_key)
        if key not in seen:
            seen[key] = item
            continue
        # Merge tags and notes when duplicates appear across chunk runs.
        existing = seen[key]
        merged_tags = sorted(set(existing.get("tags", [])) | set(item.get("tags", [])))
        existing["tags"] = merged_tags
        if not existing.get("notes") and item.get("notes"):
            existing["notes"] = item.get("notes")
    return list(seen.values())


def _extract_text_from_upload(upload: dict[str, Any]) -> str:
    content_type = upload.get("content_type") or "application/octet-stream"
    content = upload.get("content") or b""
    if content_type in ALLOWED_TEXT_TYPES:
        return content.decode("utf-8", errors="ignore")
    if content_type == PDF_TYPE:
        return _extract_pdf_text(content)
    if content_type in EXCEL_TYPES:
        return _extract_excel_text(content)
    return ""


def _clean_planner_item_line(line: str) -> str:
    without_upc = re.sub(r"\b\d{8,}\b", "", line)
    normalized = " ".join(without_upc.split())
    return normalized.strip("-: ")


def _looks_like_section_header(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if len(stripped) > 80:
        return False
    letter_chars = [ch for ch in stripped if ch.isalpha()]
    if not letter_chars:
        return False
    uppercase_ratio = sum(1 for ch in letter_chars if ch.isupper()) / len(letter_chars)
    # Planner display headers are usually short and mostly uppercase.
    return uppercase_ratio >= 0.7


def _parse_planner_sections_from_text(text: str) -> list[dict[str, str]]:
    entries: list[dict[str, str]] = []
    current_section = "Uncategorized"
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if _looks_like_section_header(line):
            current_section = line
            continue
        item_name = _clean_planner_item_line(line)
        if not item_name:
            continue
        entries.append({"item_name": item_name, "source_section": current_section, "notes": ""})
    return entries


def _normalize_text(text: str) -> str:
    return " ".join("".join(ch.lower() if ch.isalnum() else " " for ch in text).split())


def _token_set(text: str) -> set[str]:
    return {token for token in _normalize_text(text).split() if len(token) > 1}


def _match_score(submission_name: str, ad_name: str) -> float:
    normalized_submission = _normalize_text(submission_name)
    normalized_ad = _normalize_text(ad_name)
    if not normalized_submission or not normalized_ad:
        return 0.0

    if normalized_submission == normalized_ad:
        return 1.0

    ratio = difflib.SequenceMatcher(None, normalized_submission, normalized_ad).ratio()
    submission_tokens = _token_set(normalized_submission)
    ad_tokens = _token_set(normalized_ad)
    overlap = len(submission_tokens & ad_tokens)
    union = len(submission_tokens | ad_tokens)
    jaccard = (overlap / union) if union else 0.0

    contains_bonus = 0.0
    if normalized_submission in normalized_ad or normalized_ad in normalized_submission:
        contains_bonus = 0.12

    return max(ratio, jaccard + contains_bonus)


def parse_weekly_ad(uploads: list[dict[str, Any]]) -> dict[str, Any]:
    system_prompt = """
You extract structured weekly grocery ad data.
Return strict JSON only with this shape:
{
  "ad_date": "YYYY-MM-DD or empty string",
  "items": [
    {
    "name": "string",
        "size_text": "string",
    "price_text": "string",
    "page_number": 1,
    "tags": ["front_page_items", "price_lock_items", "just_4_u_items", "four_x_points_items", "five_friday_items", "member_price_items", "regular_items"],
    "notes": "string"
    }
  ]
}
Rules:
- Include one entry per ad item.
- Cover all pages from all uploaded files; do not stop after page 1.
- Do not assume an entire page belongs to a single promotion type.
- Inspect each advertised item independently.
- Use front_page_items when the item appears on page 1.
- Use just_4_u_items when the item has a Clip or Click banner, coupon-style box, or barcode shown with the offer.
- Use price_lock_items when the item has BOTH a yellow padlock icon and a blue Earn 2X Points icon.
- Use five_friday_items when the item has a red circular graphic containing a white $5.
- Use four_x_points_items when the item has a red Earn 4X Points graphic.
- Use member_price_items when none of the special indicators above are present.
- Include regular_items for general non-special ad items when promotion labels are unclear.
- Tags may contain multiple values when applicable.
- An item can belong to multiple tags at the same time.
- Include size_text whenever size, weight, count, volume, or pack information appears (for example: 12 oz, 2 lb, 24 ct, 6-pack).
- Keep names concise and normalized to what a store employee would recognize.
- Keep page_number accurate for every extracted item.
""".strip()
    instruction = "Parse these Safeway weekly ad files across every page, classify each item by visual indicators, and return the JSON structure exactly."

    def normalize_items(raw_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        for item in raw_items:
            item.setdefault("size_text", "")
            item.setdefault("price_text", "")
            item.setdefault("page_number", None)
            item.setdefault("tags", ["regular_items"])
            item.setdefault("notes", "")
            normalized.append(item)
        return normalized

    has_pdf = any((upload.get("content_type") or "") == PDF_TYPE for upload in uploads)
    if not has_pdf:
        payload = _run_json_request(system_prompt, instruction, uploads)
        items = normalize_items(payload.get("items", []))
        if not items:
            raise RoboertaAIError("No weekly ad items were extracted from the uploaded files.")
        return {"ad_date": payload.get("ad_date", ""), "items": items}

    # For PDF ads, parse page-by-page for consistent extraction coverage.
    chunked_uploads = _chunk_weekly_ad_uploads(uploads, pages_per_chunk=1)
    merged_items: list[dict[str, Any]] = []
    merged_ad_date = ""
    for chunk_index, chunk in enumerate(chunked_uploads, start=1):
        chunk_instruction = (
            f"Chunk {chunk_index} of {len(chunked_uploads)}. "
            "Parse ONLY the items visible in these files and return the JSON structure exactly."
        )
        chunk_payload = _run_json_request(system_prompt, chunk_instruction, chunk)
        if not merged_ad_date:
            merged_ad_date = chunk_payload.get("ad_date", "")
        chunk_items = normalize_items(chunk_payload.get("items", []))
        if len(chunk) == 1:
            forced_page = _extract_page_number_from_filename(chunk[0].get("filename", ""))
            if forced_page is not None:
                for item in chunk_items:
                    item["page_number"] = forced_page
        merged_items.extend(chunk_items)

    merged_items = _dedupe_ad_items(merged_items)
    if not merged_items:
        raise RoboertaAIError("No weekly ad items were extracted from the uploaded files.")

    return {
        "ad_date": merged_ad_date,
        "items": merged_items,
    }


def parse_submission(kind: str, upload: dict[str, Any]) -> dict[str, Any]:
    if kind == "planner":
        planner_text = _extract_text_from_upload(upload)
        planner_items = _parse_planner_sections_from_text(planner_text)
        if planner_items:
            return {"items": planner_items}

    source_instruction = "photo of products or tags" if kind == "photo" else "planner or display section document"
    system_prompt = f"""
You extract shopping items from a {source_instruction}.
Return strict JSON only with this shape:
{{
  "items": [
    {{
      "item_name": "string",
      "source_section": "string",
      "notes": "string"
    }}
  ]
}}
Rules:
- For planner documents, preserve the display section name in source_section when present.
- For photos, use source_section only when the image clearly shows a department or display section.
- Do not invent items that are not visible in the input.
""".strip()
    instruction = f"Extract all relevant items from this {kind} and return the JSON structure exactly."
    payload = _run_json_request(system_prompt, instruction, [upload])
    items = payload.get("items", [])
    if not items:
        raise RoboertaAIError("No items were extracted from the submitted file.")
    for item in items:
        item.setdefault("source_section", "")
        item.setdefault("notes", "")
    return payload


def compare_ad_to_submission(
    ad_date: str,
    ad_items: list[dict[str, Any]],
    submission_kind: str,
    submission_items: list[dict[str, Any]],
) -> dict[str, Any]:
    del ad_date, submission_kind

    groups: dict[str, list[dict[str, Any]]] = {group_name: [] for group_name in GROUP_ORDER}
    section_results_map: dict[str, list[dict[str, Any]]] = {}
    section_order: list[str] = []
    matched_count = 0
    unmatched_count = 0
    threshold = 0.53

    for submission in submission_items:
        submission_name = (submission.get("item_name") or "").strip()
        source_section = (submission.get("source_section") or "Uncategorized").strip() or "Uncategorized"
        if source_section not in section_results_map:
            section_results_map[source_section] = []
            section_order.append(source_section)
        if not submission_name:
            continue

        best_item: dict[str, Any] | None = None
        best_score = 0.0
        for ad_item in ad_items:
            score = _match_score(submission_name, ad_item.get("name", ""))
            if score > best_score:
                best_score = score
                best_item = ad_item

        if best_item is None or best_score < threshold:
            unmatched_count += 1
            continue

        matched_count += 1
        tags = [tag for tag in best_item.get("tags", []) if tag in GROUP_ORDER] or ["regular_items"]
        record = {
            "item_name": submission_name,
            "matched_ad_name": best_item.get("name", submission_name),
            "ad_size": best_item.get("size_text", ""),
            "ad_price": best_item.get("price_text", ""),
            "source_section": source_section,
            "notes": submission.get("notes", ""),
        }
        section_results_map[source_section].append(record)
        for tag in tags:
            groups[tag].append(record)

    highlights = [
        f"Matched {matched_count} submitted items across all ad pages.",
        f"Unmatched submitted items: {unmatched_count}.",
        "Matches are grouped by each ad item's real promo tags, not only front page.",
    ]
    section_results = [{"section_name": section_name, "matches": section_results_map.get(section_name, [])} for section_name in section_order]
    return {"highlights": highlights, "groups": groups, "section_results": section_results}
